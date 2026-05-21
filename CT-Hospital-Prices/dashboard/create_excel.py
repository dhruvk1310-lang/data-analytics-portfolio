"""
CT Hospital Price — Excel Dashboard Generator
==============================================
Reads data/processed/ outputs and builds a formatted Excel workbook:

  Sheet 1 — DASHBOARD    KPI cards + bar chart (price gap by procedure)
  Sheet 2 — BY PROCEDURE  Full comparison table, all hospitals side-by-side
  Sheet 3 — BY HOSPITAL   Each hospital's price profile + ranking
  Sheet 4 — RAW DATA      Clean long-format data for pivot tables

Usage:
    python dashboard/create_excel.py
    python dashboard/create_excel.py --demo
"""

import argparse
import json
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                              PatternFill, Side)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT     = Path(__file__).parent.parent
PROC_DIR = ROOT / "data" / "processed"
OUT_PATH = ROOT / "dashboard" / "CT_Hospital_Price_Dashboard.xlsx"

# ── Color palette ─────────────────────────────────────────────────────────────
NAVY    = "1F2D4E"
GOLD    = "C9A84C"
RED     = "C0392B"
GREEN   = "27AE60"
BLUE    = "2980B9"
LGRAY   = "F2F4F7"
MGRAY   = "BDC3C7"
WHITE   = "FFFFFF"
DKTEXT  = "1A1A2E"

HOSP_COLORS = {
    "Yale New Haven Hospital": "2980B9",
    "Hartford Hospital":       "C0392B",
    "UConn Health":            "27AE60",
    "Stamford Hospital":       "8E44AD",
    "Bridgeport Hospital":     "E67E22",
}

# ── Style helpers ─────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=DKTEXT, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def border(style="thin"):
    s = Side(style=style, color=MGRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row_num, values, widths=None,
                     bg=NAVY, fg=WHITE, size=11):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.fill  = fill(bg)
        c.font  = font(bold=True, size=size, color=fg)
        c.alignment = center()
        c.border = border()
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

def apply_data_row(ws, row_num, values, bg=WHITE, bold=False,
                   number_formats=None):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=i, value=val)
        c.fill  = fill(bg)
        c.font  = font(bold=bold)
        c.alignment = left()
        c.border = border()
        if number_formats and i <= len(number_formats) and number_formats[i-1]:
            c.number_format = number_formats[i-1]

def money_fmt(val):
    try:
        return float(val) if val and str(val) not in ("", "nan") else None
    except Exception:
        return None


# ── Sheet 1: Dashboard ────────────────────────────────────────────────────────

def build_dashboard(ws, stats: dict, comparison: pd.DataFrame,
                    hospital_names: list):
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value     = "Connecticut Hospital Price Transparency Dashboard"
    t.fill      = fill(NAVY)
    t.font      = font(bold=True, size=20, color=WHITE)
    t.alignment = center()
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:L2")
    s = ws["A2"]
    s.value     = "Gross chargemaster prices for 20 common procedures across 5 CT hospitals  |  Source: CMS 45 CFR 180 mandatory price files"
    s.fill      = fill(GOLD)
    s.font      = font(bold=False, size=10, color=WHITE)
    s.alignment = center()
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10  # spacer

    # ── KPI Cards (row 4-8, columns A-L split into 4 cards)
    kpis = [
        ("Largest\nPrice Gap",
         f"${stats['largest_dollar_gap']['gap']:,.0f}",
         f"{stats['largest_dollar_gap']['procedure']}",
         RED),
        ("Highest\nPrice Ratio",
         f"{stats['highest_price_ratio']['ratio']}×",
         f"{stats['highest_price_ratio']['procedure']}",
         NAVY),
        ("Most Expensive\nHospital",
         stats["most_expensive_hospital"].split()[0],
         "Avg gross charge",
         "8E44AD"),
        ("Cheapest\nHospital",
         stats["cheapest_hospital"].split()[0],
         "Avg gross charge",
         GREEN),
    ]

    card_cols = [("A","C"), ("D","F"), ("G","I"), ("J","L")]
    for (col_s, col_e), (label, value, sub, color) in zip(card_cols, kpis):
        ws.merge_cells(f"{col_s}4:{col_e}4")
        ws.merge_cells(f"{col_s}5:{col_e}6")
        ws.merge_cells(f"{col_s}7:{col_e}7")
        ws.merge_cells(f"{col_s}8:{col_e}8")

        lc = ws[f"{col_s}4"]
        lc.value = label; lc.fill = fill(color)
        lc.font = font(bold=True, size=10, color=WHITE); lc.alignment = center()
        ws.row_dimensions[4].height = 28

        vc = ws[f"{col_s}5"]
        vc.value = value; vc.fill = fill(color)
        vc.font = font(bold=True, size=22, color=WHITE); vc.alignment = center()
        ws.row_dimensions[5].height = 36; ws.row_dimensions[6].height = 2

        sc = ws[f"{col_s}7"]
        sc.value = sub; sc.fill = fill(color)
        sc.font = font(size=9, color="DDDDDD"); sc.alignment = center()
        ws.row_dimensions[7].height = 16; ws.row_dimensions[8].height = 10

    # ── Bar chart: top 10 procedures by dollar gap
    ws.row_dimensions[9].height = 10
    ws.merge_cells("A10:L10")
    h = ws["A10"]
    h.value = "Top Procedures by Price Gap (Most Expensive vs Cheapest Hospital)"
    h.fill  = fill(LGRAY)
    h.font  = font(bold=True, size=12)
    h.alignment = left()
    ws.row_dimensions[10].height = 28

    top10 = comparison.head(10).copy()
    chart_start_row = 11

    # Write hidden data for chart
    ws.cell(row=chart_start_row, column=1, value="Procedure")
    ws.cell(row=chart_start_row, column=2, value="Min Price")
    ws.cell(row=chart_start_row, column=3, value="Max Price")
    ws.cell(row=chart_start_row, column=4, value="Gap")

    for i, (_, row) in enumerate(top10.iterrows(), 1):
        r = chart_start_row + i
        ws.cell(row=r, column=1, value=row["label"])
        ws.cell(row=r, column=2, value=money_fmt(row["min_price"]))
        ws.cell(row=r, column=3, value=money_fmt(row["max_price"]))
        ws.cell(row=r, column=4, value=money_fmt(row["dollar_gap"]))
        # Style these rows small/hidden (white text)
        for col in range(1, 5):
            c = ws.cell(row=r, column=col)
            c.font = font(size=9, color=WHITE)
            c.fill = fill(WHITE)

    # Make header row also white/invisible
    for col in range(1, 5):
        c = ws.cell(row=chart_start_row, column=col)
        c.font = font(size=9, color=WHITE); c.fill = fill(WHITE)

    # Build chart
    chart = BarChart()
    chart.type   = "bar"  # horizontal
    chart.grouping = "clustered"
    chart.title  = None
    chart.y_axis.title = None
    chart.x_axis.title = "Price (USD)"
    chart.width  = 22
    chart.height = 14

    data_min = Reference(ws, min_col=2, min_row=chart_start_row,
                         max_row=chart_start_row + len(top10))
    data_max = Reference(ws, min_col=3, min_row=chart_start_row,
                         max_row=chart_start_row + len(top10))
    cats = Reference(ws, min_col=1, min_row=chart_start_row+1,
                     max_row=chart_start_row + len(top10))

    chart.add_data(data_min, titles_from_data=True)
    chart.add_data(data_max, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = GREEN
    chart.series[1].graphicalProperties.solidFill = RED

    ws.add_chart(chart, "A11")

    # Column widths
    for col, w in zip("ABCDEFGHIJKL", [18,8,8,8,8,8,14,8,8,14,8,8]):
        ws.column_dimensions[col].width = w


# ── Sheet 2: By Procedure ─────────────────────────────────────────────────────

def build_by_procedure(ws, comparison: pd.DataFrame, hospital_names: list):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    t = ws["A1"]
    t.value = "Price Comparison by Procedure — All Hospitals"
    t.fill  = fill(NAVY); t.font = font(bold=True, size=16, color=WHITE)
    t.alignment = center(); ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:P2")
    s = ws["A2"]
    s.value = "Gross chargemaster prices (list price before insurance). Conditional formatting: red = most expensive, green = cheapest."
    s.fill  = fill(GOLD); s.font = font(size=10, color=WHITE, italic=True)
    s.alignment = center(); ws.row_dimensions[2].height = 20

    headers = ["CPT", "Procedure", "Category"] + \
              [h.split()[0] for h in hospital_names] + \
              ["Cheapest $", "Priciest $", "Gap $", "Ratio", "Cheapest", "Priciest"]
    widths  = [8, 22, 12] + [14]*len(hospital_names) + [12,12,12,8,14,14]
    apply_header_row(ws, 3, headers, widths)

    hosp_cols_in_df = [c for c in comparison.columns if c in hospital_names]

    for row_idx, (_, row) in enumerate(comparison.iterrows(), 4):
        bg = WHITE if row_idx % 2 == 0 else LGRAY
        vals = (
            [row["cpt_code"], row["label"], row["category"]] +
            [money_fmt(row.get(h)) for h in hosp_cols_in_df] +
            [money_fmt(row["min_price"]), money_fmt(row["max_price"]),
             money_fmt(row["dollar_gap"]), money_fmt(row["price_ratio"]),
             row["cheapest"], row["most_expensive"]]
        )
        nfmt = [None, None, None] + \
               ['$#,##0']*len(hosp_cols_in_df) + \
               ['$#,##0','$#,##0','$#,##0','0.0"x"', None, None]
        apply_data_row(ws, row_idx, vals, bg=bg, number_formats=nfmt)

    # Conditional formatting on hospital price columns (cols 4 to 4+n)
    n = len(hosp_cols_in_df)
    first_data_row = 4
    last_data_row  = 3 + len(comparison)
    col_start = get_column_letter(4)
    col_end   = get_column_letter(3 + n)
    price_range = f"{col_start}{first_data_row}:{col_end}{last_data_row}"

    ws.conditional_formatting.add(
        price_range,
        ColorScaleRule(
            start_type="min",  start_color="63BE7B",   # green = cheap
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max",    end_color="F8696B",     # red = expensive
        )
    )

    # Gap column data bar
    gap_col = get_column_letter(4 + n + 2)
    ws.conditional_formatting.add(
        f"{gap_col}{first_data_row}:{gap_col}{last_data_row}",
        DataBarRule(start_type="min", start_value=0,
                    end_type="max",   end_value=None,
                    color=NAVY)
    )

    ws.freeze_panes = "A4"


# ── Sheet 3: By Hospital ──────────────────────────────────────────────────────

def build_by_hospital(ws, df_clean: pd.DataFrame, hospital_names: list,
                      rankings: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Hospital Price Profiles & Rankings"
    t.fill  = fill(NAVY); t.font = font(bold=True, size=16, color=WHITE)
    t.alignment = center(); ws.row_dimensions[1].height = 38

    # Rankings table
    ws.merge_cells("A3:H3")
    ws["A3"].value = "Overall Hospital Rankings — Average Gross Charge Across All Procedures"
    ws["A3"].fill = fill(LGRAY)
    ws["A3"].font = font(bold=True, size=12)
    ws["A3"].alignment = left()

    apply_header_row(ws, 4, ["Rank","Hospital","Avg Gross Charge","Procedures Found","vs Cheapest"],
                     [6,32,18,18,16])

    min_avg = rankings["avg_gross_charge"].min()
    for i, (_, row) in enumerate(rankings.iterrows(), 1):
        bg = "FFF0F0" if i == 1 else ("F0FFF0" if i == len(rankings) else WHITE)
        pct_above = ((row["avg_gross_charge"] - min_avg) / min_avg * 100) if min_avg else 0
        vs = f"+{pct_above:.0f}%" if i > 1 else "baseline"
        apply_data_row(ws, 4+i, [i, row["hospital"],
                                  row["avg_gross_charge"],
                                  row.get("procedures_found", "—"), vs],
                       bg=bg,
                       number_formats=[None,None,"$#,##0",None,None])

    # Per-hospital breakdown below
    current_row = 4 + len(rankings) + 3
    for hosp in hospital_names:
        color = HOSP_COLORS.get(hosp, NAVY)
        sub = df_clean[df_clean["hospital"] == hosp].copy()
        if sub.empty:
            continue

        ws.merge_cells(f"A{current_row}:H{current_row}")
        h = ws[f"A{current_row}"]
        h.value = hosp
        h.fill  = fill(color); h.font = font(bold=True, size=13, color=WHITE)
        h.alignment = left(); ws.row_dimensions[current_row].height = 30
        current_row += 1

        apply_header_row(ws, current_row,
                         ["CPT","Procedure","Gross Charge","Cash Price","Min Neg.","Max Neg."],
                         [8,28,14,13,12,12], bg=color)
        current_row += 1

        sub = sub.sort_values("gross_charge", ascending=False)
        for _, row in sub.iterrows():
            label = TARGET_CPTS.get(str(row["cpt_code"]), (row["cpt_code"],))[0] \
                    if "TARGET_CPTS" in dir() else row.get("description","")[:30]
            bg = WHITE if current_row % 2 == 0 else LGRAY
            apply_data_row(ws, current_row,
                           [row["cpt_code"], row.get("description","")[:35],
                            money_fmt(row["gross_charge"]),
                            money_fmt(row["discounted_cash"]),
                            money_fmt(row["min_negotiated"]),
                            money_fmt(row["max_negotiated"])],
                           bg=bg,
                           number_formats=[None,None,"$#,##0","$#,##0","$#,##0","$#,##0"])
            current_row += 1

        current_row += 2  # spacer

    for col, w in zip("ABCDEFGH", [8,32,16,14,12,12,14,14]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A5"


# ── Sheet 4: Raw Data ─────────────────────────────────────────────────────────

def build_raw_data(ws, df_clean: pd.DataFrame):
    ws.sheet_view.showGridLines = True

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Raw Data — Use this sheet to build Pivot Tables"
    t.fill  = fill(NAVY); t.font = font(bold=True, size=13, color=WHITE)
    t.alignment = center(); ws.row_dimensions[1].height = 30

    headers = ["Hospital","CPT Code","Description","Setting",
               "Gross Charge","Discounted Cash","Min Negotiated","Max Negotiated"]
    apply_header_row(ws, 2, headers,
                     [28,10,35,12,14,16,14,14])

    for row_idx, (_, row) in enumerate(df_clean.iterrows(), 3):
        bg = WHITE if row_idx % 2 == 0 else LGRAY
        apply_data_row(ws, row_idx,
                       [row["hospital"], row["cpt_code"],
                        str(row.get("description",""))[:40],
                        row.get("setting",""),
                        money_fmt(row["gross_charge"]),
                        money_fmt(row["discounted_cash"]),
                        money_fmt(row["min_negotiated"]),
                        money_fmt(row["max_negotiated"])],
                       bg=bg,
                       number_formats=[None,None,None,None,
                                       "$#,##0","$#,##0","$#,##0","$#,##0"])

    ws.auto_filter.ref = f"A2:H{2 + len(df_clean)}"
    ws.freeze_panes = "A3"


# ── Main ──────────────────────────────────────────────────────────────────────

def main(demo: bool = False):
    json_path = PROC_DIR / "summary_stats.json"
    csv_comp  = PROC_DIR / "price_comparison.csv"
    csv_clean = PROC_DIR / "all_prices_clean.csv"

    if not json_path.exists():
        print("⚠  Run analysis/analyze.py first.")
        sys.exit(1)

    with open(json_path) as f:
        stats_data = json.load(f)

    comparison  = pd.read_csv(csv_comp)
    df_clean    = pd.read_csv(csv_clean)
    rankings    = pd.DataFrame(stats_data["rankings"])
    hosp_names  = stats_data["hospital_names"]
    stats       = stats_data["stats"]

    print("Building Excel dashboard...")
    wb = Workbook()

    # Sheet 1 — Dashboard
    ws1 = wb.active
    ws1.title = "Dashboard"
    build_dashboard(ws1, stats, comparison, hosp_names)
    print("  ✓ Dashboard sheet")

    # Sheet 2 — By Procedure
    ws2 = wb.create_sheet("By Procedure")
    build_by_procedure(ws2, comparison, hosp_names)
    print("  ✓ By Procedure sheet")

    # Sheet 3 — By Hospital
    ws3 = wb.create_sheet("By Hospital")
    sys.path.insert(0, str(Path(__file__).parent.parent / "scraper"))
    try:
        from hospitals import TARGET_CPTS as _tc
        import builtins
        builtins.TARGET_CPTS = _tc
    except Exception:
        pass
    build_by_hospital(ws3, df_clean, hosp_names, rankings)
    print("  ✓ By Hospital sheet")

    # Sheet 4 — Raw Data
    ws4 = wb.create_sheet("Raw Data")
    build_raw_data(ws4, df_clean)
    print("  ✓ Raw Data sheet")

    # Tab colors
    ws1.sheet_properties.tabColor = NAVY
    ws2.sheet_properties.tabColor = BLUE
    ws3.sheet_properties.tabColor = GREEN
    ws4.sheet_properties.tabColor = MGRAY

    wb.save(OUT_PATH)
    print(f"\n✓ Dashboard saved → {OUT_PATH}")

    # Auto-open
    try:
        subprocess.run(["open", str(OUT_PATH)], check=True)
        print("  Opening in Excel...")
    except Exception:
        print(f"  Open manually: {OUT_PATH}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--demo", action="store_true")
    args = parser.parse_args()
    main(demo=args.demo)
